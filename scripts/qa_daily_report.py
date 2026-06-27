#!/usr/bin/env python3
"""
飞书群每日 QA 客观执行差异播报 (GitHub Actions 版)
读取仓库中的 Excel，计算 TOP 差异人员和问题项，推送到飞书群
"""

import json
import os
import sys
from datetime import datetime
from collections import defaultdict

import openpyxl
import requests

# ── 仓库内路径 ──────────────────────────────────────────────
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(REPO_ROOT, "data", "客观执行问题.xlsx")


def read_excel(file_path):
    """读取 Excel，返回数据行列表"""
    if not os.path.exists(file_path):
        print(f"[ERROR] Excel 文件不存在: {file_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    print(f"[INFO] Sheet: {ws.title}, {ws.max_row} 行")

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        rows.append({
            "date": str(row[0])[:10] if row[0] else "",
            "site": row[1] or "",
            "person": row[3] or "",
            "main_cat": row[6] or "",
            "prob_cat": row[8] or "",
            "l1_item": row[9] or "",
            "new_l3": row[14] or "",
            "is_diff": row[21],
            "diff_count": row[22] or 0,
        })
    print(f"[INFO] 有效数据行: {len(rows)}")
    return rows


def compute_stats(rows):
    """计算差异统计"""
    exec_rows = [r for r in rows if r["main_cat"] == "客观问题" and r["prob_cat"] == "执行问题"]

    if not exec_rows:
        print("[WARN] 没有客观+执行问题数据")
        return None

    person_diff = defaultdict(int)
    issue_diff = defaultdict(int)
    site_diff = defaultdict(int)
    dates = set()

    for r in exec_rows:
        person_diff[r["person"]] += r["diff_count"]
        label = f"{r['l1_item']} — {r['new_l3']}" if r["new_l3"] and r["new_l3"] != "-" else r["l1_item"]
        issue_diff[label] += r["diff_count"]
        site_diff[r["site"]] += r["diff_count"]
        if r["date"]:
            dates.add(r["date"])

    return {
        "total": sum(person_diff.values()),
        "date_range": f"{min(dates)} ~ {max(dates)}" if dates else "无数据",
        "sites": dict(site_diff),
        "top_persons": sorted(person_diff.items(), key=lambda x: -x[1]),
        "top_issues": sorted(issue_diff.items(), key=lambda x: -x[1]),
    }


def build_card(stats, top_n=5):
    """构造飞书交互卡片"""
    today = datetime.now().strftime("%Y-%m-%d")

    site_list = "、".join([f"{k}: {v}条" for k, v in stats["sites"].items()])

    person_lines = []
    for i, (name, count) in enumerate(stats["top_persons"][:top_n], 1):
        medal = ["🥇", "🥈", "🥉"][i - 1] if i <= 3 else f"{i}."
        person_lines.append(f"{medal} **{name}**　　{count} 条差异")

    issue_lines = []
    for i, (name, count) in enumerate(stats["top_issues"][:top_n], 1):
        medal = ["🏆", "⚠️", "⚠️"][i - 1] if i <= 3 else "▸"
        issue_lines.append(f"{medal} **{name}**　　{count} 条差异")

    card_md = (
        f"**📅 统计周期：{stats['date_range']}**\n"
        f"**📍 覆盖站点：{site_list}**\n\n---\n\n"
        f"**👤 TOP{top_n} 差异人员**\n{chr(10).join(person_lines)}\n\n---\n\n"
        f"**⚠️ TOP{top_n} 差异问题项**\n{chr(10).join(issue_lines)}"
    )

    return {
        "msg_type": "interactive",
        "card": {
            "header": {
                "title": {"tag": "plain_text", "content": f"📱 手机QA客观执行差异日报 | {today}"},
                "template": "blue",
            },
            "elements": [
                {"tag": "div", "text": {"tag": "lark_md", "content": card_md}},
                {"tag": "hr"},
                {
                    "tag": "note",
                    "elements": [
                        {"tag": "plain_text", "content": f"📊 客观执行差异共 {stats['total']} 条　|　自动播报 @ {today}"}
                    ],
                },
            ],
        },
    }


def send_to_feishu(webhook_url, card):
    """发送卡片到飞书"""
    print(f"[INFO] 发送到飞书...")
    try:
        resp = requests.post(webhook_url, json=card, timeout=15)
        resp.raise_for_status()
        result = resp.json()
        print(f"[INFO] 响应: {json.dumps(result, ensure_ascii=False)}")
        if result.get("code") == 0 or result.get("StatusCode") == 0:
            print("[OK] 发送成功！")
            return True
        else:
            print(f"[ERROR] 飞书错误: {result}")
            return False
    except requests.exceptions.RequestException as e:
        print(f"[ERROR] 请求失败: {e}")
        return False


def main():
    print(f"[INFO] ====== QA 客观执行差异日报 ======")
    print(f"[INFO] 时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    webhook_url = os.environ.get("FEISHU_WEBHOOK_URL", "")
    if not webhook_url:
        print("[ERROR] 未设置 FEISHU_WEBHOOK_URL 环境变量")
        sys.exit(1)

    rows = read_excel(EXCEL_PATH)
    stats = compute_stats(rows)

    if stats is None:
        print("[WARN] 无差异数据，跳过")
        return

    print(f"[INFO] 执行差异: {stats['total']} 条")
    print(f"[INFO] TOP人员: {[p[0] for p in stats['top_persons'][:5]]}")
    print(f"[INFO] TOP问题: {[p[0] for p in stats['top_issues'][:5]]}")

    card = build_card(stats)
    success = send_to_feishu(webhook_url, card)
    if not success:
        sys.exit(1)


if __name__ == "__main__":
    main()
